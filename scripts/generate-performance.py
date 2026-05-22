try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from copy import copy
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import json
import sys
import os
import re
import requests
import urllib.parse
from datetime import datetime
from collections import defaultdict, OrderedDict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
CONFIG_PATH = os.path.join(PROJECT_ROOT, 'config.json')
TEMPLATE_PATH = os.path.join(PROJECT_ROOT, 'template', 'template.xlsx')

BUG_BACKLOG_TYPE_IDS = {'6966f89927d8ec1063e68c33', '6a06e3ce5919695071aa9619'}

HEADER_KEYWORDS = {
    'project': ['项目名称', '项目', '项目名', '项目标识', '所属项目', '项目名'],
    'date': ['日期', '时间', '工作日期', '完成日期', '工作时间', '日期范围'],
    'content': ['工作内容', '内容', '工作完成情况', '完成情况', '工作描述', '工作事项', '工作'],
}


def analyze_template_columns(template_path):
    if not HAS_OPENPYXL or not os.path.exists(template_path):
        return {0: 'project', 1: 'date', 2: 'content'}

    wb = openpyxl.load_workbook(template_path)
    ws = None
    for sheet in wb.worksheets:
        if '工作完成情况' in sheet.title or '完成' in sheet.title:
            ws = sheet
            break
    if ws is None:
        ws = wb.active

    if ws.max_row < 1 or ws.max_column < 1:
        wb.close()
        return {0: 'project', 1: 'date', 2: 'content'}

    first_row_vals = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            first_row_vals.append(str(val).strip())

    header_mapping = _try_header_match(first_row_vals, ws.max_column)
    if header_mapping:
        wb.close()
        return header_mapping

    column_values = {}
    for col in range(1, ws.max_column + 1):
        values = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip():
                values.append(str(val).strip())
        column_values[col - 1] = values

    wb.close()

    if not column_values:
        return {0: 'project', 1: 'date', 2: 'content'}

    mapping = _score_columns(column_values)
    return mapping


def _try_header_match(first_row_vals, max_col):
    if not first_row_vals:
        return None

    mapping = {}
    used_semantics = set()

    for col_idx, val in enumerate(first_row_vals):
        for semantic, keywords in HEADER_KEYWORDS.items():
            if semantic in used_semantics:
                continue
            for kw in keywords:
                if kw in val:
                    mapping[col_idx] = semantic
                    used_semantics.add(semantic)
                    break

    if len(used_semantics) >= 2:
        for col_idx in range(max_col):
            if col_idx not in mapping:
                for semantic in ['project', 'date', 'content']:
                    if semantic not in used_semantics:
                        mapping[col_idx] = semantic
                        used_semantics.add(semantic)
                        break
        return mapping

    return None


def _score_columns(column_values):
    col_scores = {}
    for col_idx, values in column_values.items():
        date_score = 0
        project_score = 0
        content_score = 0

        for val in values:
            if re.match(r'^\d{1,2}\.\d{1,2}(-\d{1,2}\.\d{1,2})?$', val):
                date_score += 3
            elif re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$', val):
                date_score += 3

            if len(val) <= 15:
                project_score += 1
            else:
                project_score -= 0.5

            if len(val) > 30:
                content_score += 2
            if re.search(r'[1-9][、.．]', val):
                content_score += 2
            if '完成' in val or '解决' in val or '进行中' in val:
                content_score += 1

        col_scores[col_idx] = {
            'date': date_score,
            'project': project_score,
            'content': content_score
        }

    mapping = {}
    used_semantics = set()

    best_date_col = max(col_scores, key=lambda k: col_scores[k]['date'])
    if col_scores[best_date_col]['date'] > 0:
        mapping[best_date_col] = 'date'
        used_semantics.add('date')

    remaining = {k: v for k, v in col_scores.items() if k not in mapping}
    if remaining:
        best_content_col = max(remaining, key=lambda k: remaining[k]['content'])
        if remaining[best_content_col]['content'] > 0:
            mapping[best_content_col] = 'content'
            used_semantics.add('content')

    remaining = {k: v for k, v in col_scores.items() if k not in mapping}
    if remaining:
        best_project_col = max(remaining, key=lambda k: remaining[k]['project'])
        mapping[best_project_col] = 'project'
        used_semantics.add('project')

    default_mapping = {0: 'project', 1: 'date', 2: 'content'}
    for col_idx in column_values:
        if col_idx not in mapping:
            mapping[col_idx] = default_mapping.get(col_idx, 'content')

    for semantic in ['project', 'date', 'content']:
        if semantic not in set(mapping.values()):
            for col_idx in range(3):
                if col_idx not in mapping:
                    mapping[col_idx] = semantic
                    break

    return mapping


def read_config():
    if not os.path.exists(CONFIG_PATH):
        print(f'ERROR: config.json not found at {CONFIG_PATH}')
        sys.exit(1)
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_access_token(config):
    api_config = config.get("api", {})
    base_url = api_config.get("base_url", "https://open.pingcode.com")
    params = {
        "grant_type": api_config.get("grant_type", "client_credentials"),
        "client_id": api_config.get("client_id"),
        "client_secret": api_config.get("client_secret")
    }
    query_string = urllib.parse.urlencode(params)
    url = f"{base_url}/v1/auth/token?{query_string}"
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        return response.json().get("access_token")
    except Exception as e:
        print(f"ERROR: Failed to get access token: {e}")
        return None


def is_bug(item):
    if item.get('type') == 'bug':
        return True
    if item.get('type') == 'story':
        props = item.get('properties', {}) or {}
        backlog_type = props.get('backlog_type')
        if backlog_type in BUG_BACKLOG_TYPE_IDS:
            return True
    return False


def fetch_all_work_items(config, token, project_id):
    base_url = config.get("api", {}).get("base_url", "https://open.pingcode.com")
    url = f"{base_url}/v1/project/work_items"
    headers = {"Authorization": f"Bearer {token}"}

    all_items = []
    seen_ids = set()

    for item_type in ['story', 'task', 'bug']:
        page = 0
        while True:
            params = {
                "project_id": project_id,
                "type": item_type,
                "page_size": 100,
                "page_index": page
            }
            try:
                response = requests.get(url, headers=headers, params=params, timeout=30)
                if response.status_code != 200:
                    break
                data = response.json()
                values = data.get("values", [])
                for item in values:
                    item_id = item.get('id')
                    if item_id and item_id not in seen_ids:
                        all_items.append(item)
                        seen_ids.add(item_id)
                total = data.get("total", 0)
                if len(values) == 0 or (page + 1) * 100 >= total:
                    break
                page += 1
            except Exception as e:
                print(f"ERROR fetching {item_type} page {page}: {e}", file=sys.stderr)
                break

    return all_items


def filter_by_month(items, year, month):
    start_ts = int(datetime(year, month, 1).timestamp())
    if month == 12:
        end_ts = int(datetime(year + 1, 1, 1).timestamp())
    else:
        end_ts = int(datetime(year, month + 1, 1).timestamp())

    filtered = []
    for item in items:
        updated_at = item.get('updated_at') or item.get('created_at')
        if updated_at and start_ts <= updated_at < end_ts:
            filtered.append(item)
    return filtered


def build_task_bug_map(all_items):
    task_map = {}
    for item in all_items:
        if item.get('type') != 'task':
            continue
        item_title = item.get('title', '')
        if not (item_title.startswith('\u6392\u67e5\uff1a') or item_title.startswith('\u89e3\u51b3\uff1a')):
            continue
        a = item.get('assignee', {}) or {}
        name = a.get('display_name', '').strip()
        if item_title not in task_map:
            task_map[item_title] = []
        if name and name not in task_map[item_title]:
            task_map[item_title].append(name)
    return task_map


def find_bug_assignees(bug_item, task_map):
    bug_title = bug_item.get('title', '')
    investigators = []
    resolvers = []
    for task_title, names in task_map.items():
        if task_title.startswith('\u6392\u67e5\uff1a'):
            core = task_title[3:]
            if core in bug_title or bug_title in core:
                for name in names:
                    if name and name not in investigators:
                        investigators.append(name)
        elif task_title.startswith('\u89e3\u51b3\uff1a'):
            core = task_title[3:]
            if core in bug_title or bug_title in core:
                for name in names:
                    if name and name not in resolvers:
                        resolvers.append(name)
    return '\u3001'.join(investigators), '\u3001'.join(resolvers)


def format_work_item(item, task_map):
    state = item.get('state', {}) or {}
    assignee = item.get('assignee', {}) or {}
    sprint = item.get('sprint', {}) or {}
    props = item.get('properties', {}) or {}

    created_at = item.get('created_at')
    completed_at = item.get('completed_at')
    updated_at = item.get('updated_at')

    if completed_at and state.get('type') == 'completed':
        work_date = completed_at
    elif updated_at:
        work_date = updated_at
    elif created_at:
        work_date = created_at
    else:
        work_date = None

    hours = None
    if created_at and completed_at:
        hours = round((completed_at - created_at) / 3600, 1)

    module = props.get('caidanmokuai', '') or ''
    sprint_name = sprint.get('name', '') or ''

    title = item.get('title', '')
    state_type = state.get('type', '')
    item_type = item.get('type', '')
    item_is_bug = is_bug(item)

    assignee_name = ''
    if item_is_bug:
        inv, res = find_bug_assignees(item, task_map)
        assignee_name = res or inv or assignee.get('display_name', '')
    else:
        assignee_name = assignee.get('display_name', '')

    return {
        'id': item.get('identifier', ''),
        'title': title,
        'type': item_type,
        'state': state.get('name', ''),
        'state_type': state_type,
        'assignee': assignee_name,
        'sprint': sprint_name,
        'module': module,
        'created': datetime.fromtimestamp(created_at).strftime('%Y-%m-%d') if created_at else '',
        'completed': datetime.fromtimestamp(completed_at).strftime('%Y-%m-%d') if completed_at else '',
        'work_date': datetime.fromtimestamp(work_date).strftime('%Y-%m-%d') if work_date else '',
        'hours': hours,
        'is_bug': item_is_bug
    }


def group_by_assignee_project_date(items, project_name=None):
    by_assignee = defaultdict(list)
    for item in items:
        name = item['assignee'] or '\u672a\u5206\u914d'
        if not name or name == 'Ping':
            continue
        for person in name.replace('\u3001', ',').split(','):
            person = person.strip()
            if person and person != 'Ping':
                by_assignee[person].append(item)

    result = OrderedDict()
    for person, person_items in by_assignee.items():
        by_project = OrderedDict()
        for item in person_items:
            if project_name:
                proj = project_name
            else:
                proj = item['module'] or item['sprint'] or '\u5176\u4ed6'
            if proj not in by_project:
                by_project[proj] = []
            by_project[proj].append(item)

        sorted_projects = OrderedDict(
            sorted(by_project.items(), key=lambda x: -len(x[1]))
        )

        person_data = OrderedDict()
        for project, project_items in sorted_projects.items():
            by_date = defaultdict(list)
            for item in project_items:
                date_str = item['work_date']
                if date_str:
                    by_date[date_str].append(item)
                else:
                    by_date['9999-12-31'].append(item)

            sorted_dates = OrderedDict(
                sorted(by_date.items(), key=lambda x: x[0])
            )
            person_data[project] = sorted_dates

        result[person] = person_data

    return result


def format_date_cell(date_str):
    if not date_str or date_str == '9999-12-31':
        return ''
    d = datetime.strptime(date_str, '%Y-%m-%d')
    return f'{d.month}.{d.day}'


def format_work_content(items):
    lines = []
    idx = 1

    completed = [i for i in items if i['state_type'] == 'completed']
    in_progress = [i for i in items if i['state_type'] == 'in_progress']
    others = [i for i in items if i['state_type'] not in ('completed', 'in_progress')]

    for item in completed:
        title = item['title']
        if item['is_bug']:
            if title.startswith('\u6392\u67e5\uff1a') or title.startswith('\u6392\u67e5:'):
                lines.append(f'{idx}\u3001{title}')
            elif title.startswith('\u89e3\u51b3\uff1a') or title.startswith('\u89e3\u51b3:'):
                lines.append(f'{idx}\u3001{title}')
            else:
                lines.append(f'{idx}\u3001\u89e3\u51b3\uff1a{title}')
        else:
            lines.append(f'{idx}\u3001\u5b8c\u6210\uff1a{title}')
        idx += 1

    for item in in_progress:
        lines.append(f'{idx}\u3001\u8fdb\u884c\u4e2d\uff1a{item["title"]}')
        idx += 1

    for item in others:
        lines.append(f'{idx}\u3001{item["title"]}')
        idx += 1

    return '\n'.join(lines)


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def generate_excel(template_path, output_path, grouped_data, year, month):
    if not HAS_OPENPYXL:
        print("WARNING: openpyxl not installed, generating CSV instead")
        generate_csv(output_path, grouped_data, year, month)
        return True

    col_mapping = analyze_template_columns(template_path)
    reverse_map = {v: k + 1 for k, v in col_mapping.items()}
    project_col = reverse_map.get('project', 1)
    date_col = reverse_map.get('date', 2)
    content_col = reverse_map.get('content', 3)
    total_cols = max(project_col, date_col, content_col)

    print(f"Template column mapping: project→col {project_col}, date→col {date_col}, content→col {content_col}", file=sys.stderr)

    template_wb = None
    style_cache = {}
    col_widths = {}

    if os.path.exists(template_path):
        template_wb = openpyxl.load_workbook(template_path)
        for template_ws in template_wb.worksheets:
            for col in range(1, total_cols + 1):
                if col not in style_cache:
                    style_cache[col] = template_ws.cell(row=1, column=col)
            for col in range(1, template_ws.max_column + 1):
                col_letter = get_column_letter(col)
                if col_letter not in col_widths and template_ws.column_dimensions[col_letter].width:
                    col_widths[col_letter] = template_ws.column_dimensions[col_letter].width

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    default_font = Font(size=14, color='FF000000')
    default_border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    default_fill = PatternFill(patternType='solid', fgColor='FFFFFFFF')
    default_alignment = Alignment(vertical='center', wrap_text=True)
    content_alignment = Alignment(vertical='top', wrap_text=True)

    for person_name, projects in grouped_data.items():
        sheet_name = person_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if col_widths:
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[get_column_letter(project_col)].width = 12
            ws.column_dimensions[get_column_letter(date_col)].width = 14
            ws.column_dimensions[get_column_letter(content_col)].width = 80

        row = 1
        for project_name, dates in projects.items():
            project_start_row = row
            date_count = len(dates)

            for date_str, items in dates.items():
                date_cell = format_date_cell(date_str)
                content = format_work_content(items)

                if row == project_start_row:
                    ws.cell(row=row, column=project_col, value=project_name)
                else:
                    ws.cell(row=row, column=project_col, value='')

                ws.cell(row=row, column=date_col, value=date_cell)
                ws.cell(row=row, column=content_col, value=content)

                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    if col in style_cache:
                        copy_cell_style(style_cache[col], cell)
                    else:
                        cell.font = default_font
                        cell.border = default_border
                        cell.fill = default_fill
                    if col == project_col:
                        cell.alignment = default_alignment
                    elif col == date_col:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif col == content_col:
                        cell.alignment = content_alignment

                row += 1

            if date_count > 1:
                ws.merge_cells(
                    start_row=project_start_row,
                    start_column=project_col,
                    end_row=project_start_row + date_count - 1,
                    end_column=project_col
                )

    if template_wb:
        template_wb.close()

    if default_sheet and default_sheet.title in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[default_sheet.title]

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f'SUCCESS: Performance Excel saved to {output_path}')
    return True


def generate_csv(output_path, grouped_data, year, month):
    lines = []
    for person_name, projects in grouped_data.items():
        lines.append(f'=== {person_name} ===')
        for project_name, dates in projects.items():
            for date_str, items in dates.items():
                date_cell = format_date_cell(date_str)
                content = format_work_content(items)
                lines.append(f'{project_name}\t{date_cell}\t{content}')
        lines.append('')

    csv_path = output_path.replace('.xlsx', '.csv')
    with open(csv_path, 'w', encoding='utf-8-sig') as f:
        f.write('\n'.join(lines))
    print(f'SUCCESS: Performance CSV saved to {csv_path}')


def main():
    config = read_config()

    year = datetime.now().year
    month = datetime.now().month
    project_id = None
    project_name = None
    output_path = None
    assignee = None

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--year' and i + 1 < len(sys.argv):
            year = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--month' and i + 1 < len(sys.argv):
            month = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--project-id' and i + 1 < len(sys.argv):
            project_id = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--project-name' and i + 1 < len(sys.argv):
            project_name = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--output' and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--assignee' and i + 1 < len(sys.argv):
            assignee = sys.argv[i + 1]
            i += 2
        else:
            i += 1

    token = get_access_token(config)
    if not token:
        print("ERROR: Failed to get access token")
        sys.exit(1)

    if not project_id:
        project_id = config.get("project", {}).get("default_project_id")
    if not project_id:
        print("ERROR: project_id not specified. Use --project-id or set default_project_id in config")
        sys.exit(1)

    print(f"Fetching work items for project {project_id}, {year}-{month}...", file=sys.stderr)
    all_items = fetch_all_work_items(config, token, project_id)
    print(f"Total items fetched: {len(all_items)}", file=sys.stderr)

    monthly_items = filter_by_month(all_items, year, month)
    print(f"Items in {year}-{month}: {len(monthly_items)}", file=sys.stderr)

    if not monthly_items:
        print("NO_DATA:No work items found for the given period")
        sys.exit(0)

    task_map = build_task_bug_map(all_items)

    formatted = [format_work_item(item, task_map) for item in monthly_items]

    grouped = group_by_assignee_project_date(formatted, project_name=project_name)

    if assignee:
        filtered = OrderedDict()
        for person, projects in grouped.items():
            if assignee in person or person in assignee:
                filtered[person] = projects
        grouped = filtered

    if not grouped:
        print("NO_DATA:No performance data found for the given period")
        sys.exit(0)

    if not output_path:
        output_dir = os.path.join(PROJECT_ROOT, 'output')
        output_path = os.path.join(output_dir, f'performance-{year}{month:02d}.xlsx')

    if not os.path.isabs(output_path):
        output_path = os.path.abspath(output_path)

    generate_excel(TEMPLATE_PATH, output_path, grouped, year, month)


if __name__ == '__main__':
    main()
