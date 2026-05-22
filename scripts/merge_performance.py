import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
from collections import OrderedDict
import sys
import os
import re
import glob

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'output')
TEMPLATE_PATH = os.path.join(PROJECT_ROOT, 'template', 'template.xlsx')

HEADER_KEYWORDS = {
    'project': ['项目名称', '项目', '项目名', '项目标识', '所属项目', '项目名'],
    'date': ['日期', '时间', '工作日期', '完成日期', '工作时间', '日期范围'],
    'content': ['工作内容', '内容', '工作完成情况', '完成情况', '工作描述', '工作事项', '工作'],
}


def analyze_template_columns(template_path):
    if not os.path.exists(template_path):
        return {0: 'project', 1: 'date', 2: 'content'}

    wb = openpyxl.load_workbook(template_path)
    ws = None
    for sheet in wb.worksheets:
        if '工作完成情况' in sheet.title or '完成' in sheet.title:
            ws = sheet
            break
    if ws is None:
        ws = wb.active

    if ws.max_row < 1 or ws.max_column < 1:
        wb.close()
        return {0: 'project', 1: 'date', 2: 'content'}

    first_row_vals = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            first_row_vals.append(str(val).strip())

    header_mapping = _try_header_match(first_row_vals, ws.max_column)
    if header_mapping:
        wb.close()
        return header_mapping

    column_values = {}
    for col in range(1, ws.max_column + 1):
        values = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip():
                values.append(str(val).strip())
        column_values[col - 1] = values

    wb.close()

    if not column_values:
        return {0: 'project', 1: 'date', 2: 'content'}

    mapping = _score_columns(column_values)
    return mapping


def _try_header_match(first_row_vals, max_col):
    if not first_row_vals:
        return None

    mapping = {}
    used_semantics = set()

    for col_idx, val in enumerate(first_row_vals):
        for semantic, keywords in HEADER_KEYWORDS.items():
            if semantic in used_semantics:
                continue
            for kw in keywords:
                if kw in val:
                    mapping[col_idx] = semantic
                    used_semantics.add(semantic)
                    break

    if len(used_semantics) >= 2:
        for col_idx in range(max_col):
            if col_idx not in mapping:
                for semantic in ['project', 'date', 'content']:
                    if semantic not in used_semantics:
                        mapping[col_idx] = semantic
                        used_semantics.add(semantic)
                        break
        return mapping

    return None


def _score_columns(column_values):
    col_scores = {}
    for col_idx, values in column_values.items():
        date_score = 0
        project_score = 0
        content_score = 0

        for val in values:
            if re.match(r'^\d{1,2}\.\d{1,2}(-\d{1,2}\.\d{1,2})?$', val):
                date_score += 3
            elif re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$', val):
                date_score += 3

            if len(val) <= 15:
                project_score += 1
            else:
                project_score -= 0.5

            if len(val) > 30:
                content_score += 2
            if re.search(r'[1-9][、.．]', val):
                content_score += 2
            if '完成' in val or '解决' in val or '进行中' in val:
                content_score += 1

        col_scores[col_idx] = {
            'date': date_score,
            'project': project_score,
            'content': content_score
        }

    mapping = {}
    used_semantics = set()

    best_date_col = max(col_scores, key=lambda k: col_scores[k]['date'])
    if col_scores[best_date_col]['date'] > 0:
        mapping[best_date_col] = 'date'
        used_semantics.add('date')

    remaining = {k: v for k, v in col_scores.items() if k not in mapping}
    if remaining:
        best_content_col = max(remaining, key=lambda k: remaining[k]['content'])
        if remaining[best_content_col]['content'] > 0:
            mapping[best_content_col] = 'content'
            used_semantics.add('content')

    remaining = {k: v for k, v in col_scores.items() if k not in mapping}
    if remaining:
        best_project_col = max(remaining, key=lambda k: remaining[k]['project'])
        mapping[best_project_col] = 'project'
        used_semantics.add('project')

    default_mapping = {0: 'project', 1: 'date', 2: 'content'}
    for col_idx in column_values:
        if col_idx not in mapping:
            mapping[col_idx] = default_mapping.get(col_idx, 'content')

    for semantic in ['project', 'date', 'content']:
        if semantic not in set(mapping.values()):
            for col_idx in range(3):
                if col_idx not in mapping:
                    mapping[col_idx] = semantic
                    break

    return mapping


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def load_template_styles():
    col_mapping = analyze_template_columns(TEMPLATE_PATH)
    reverse_map = {v: k + 1 for k, v in col_mapping.items()}
    project_col = reverse_map.get('project', 1)
    date_col = reverse_map.get('date', 2)
    content_col = reverse_map.get('content', 3)
    total_cols = max(project_col, date_col, content_col)

    style_cache = {}
    col_widths = {}
    template_wb = None

    if os.path.exists(TEMPLATE_PATH):
        template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
        for template_ws in template_wb.worksheets:
            for col in range(1, total_cols + 1):
                if col not in style_cache:
                    style_cache[col] = template_ws.cell(row=1, column=col)
            for col in range(1, template_ws.max_column + 1):
                col_letter = get_column_letter(col)
                if col_letter not in col_widths and template_ws.column_dimensions[col_letter].width:
                    col_widths[col_letter] = template_ws.column_dimensions[col_letter].width

    return style_cache, col_widths, template_wb, project_col, date_col, content_col


def read_person_data(filepath, project_col, date_col, content_col):
    wb = openpyxl.load_workbook(filepath)
    person_data = OrderedDict()
    col_indices = sorted(set([project_col, date_col, content_col]))
    max_col = max(col_indices)

    for ws in wb.worksheets:
        name = ws.title
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=False):
            vals = {}
            for ci in col_indices:
                vals[ci] = row[ci - 1].value if ci - 1 < len(row) else ''
                vals[ci] = vals[ci] if vals[ci] is not None else ''

            proj_val = vals.get(project_col, '')
            date_val = vals.get(date_col, '')
            content_val = vals.get(content_col, '')

            if not proj_val and not date_val and not content_val:
                continue
            rows.append((proj_val, date_val, content_val))
        if rows:
            person_data[name] = rows
    wb.close()
    return person_data


def list_all_persons(project_col, date_col, content_col):
    all_persons = OrderedDict()
    pattern = os.path.join(OUTPUT_DIR, 'performance-*-*.xlsx')

    for filepath in sorted(glob.glob(pattern)):
        basename = os.path.basename(filepath)
        if basename.startswith('performance-') and len(basename.split('-')) >= 3:
            person_data = read_person_data(filepath, project_col, date_col, content_col)
            for name, rows in person_data.items():
                if name not in all_persons:
                    all_persons[name] = []
                all_persons[name].extend(rows)

    return all_persons


def generate_for_assignee(assignee_name, output_path):
    style_cache, col_widths, template_wb, project_col, date_col, content_col = load_template_styles()
    total_cols = max(project_col, date_col, content_col)

    print(f"Template column mapping: project→col {project_col}, date→col {date_col}, content→col {content_col}", file=sys.stderr)

    fallback_font = Font(size=14, color='FF000000')
    fallback_border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    fallback_fill = PatternFill(patternType='solid', fgColor='FFFFFFFF')

    all_persons = list_all_persons(project_col, date_col, content_col)

    matched_persons = OrderedDict()
    for name, rows in all_persons.items():
        if assignee_name in name or name in assignee_name:
            matched_persons[name] = rows

    if not matched_persons:
        print(f'NO_DATA:No data found for assignee "{assignee_name}"')
        return False

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    for person_name, rows in matched_persons.items():
        sheet_name = person_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if col_widths:
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[get_column_letter(project_col)].width = 12
            ws.column_dimensions[get_column_letter(date_col)].width = 14
            ws.column_dimensions[get_column_letter(content_col)].width = 80

        row_idx = 1
        current_project = None
        project_start_row = None

        for proj_val, date_val, content_val in rows:
            if proj_val:
                if current_project is not None and project_start_row is not None and project_start_row < row_idx - 1:
                    ws.merge_cells(
                        start_row=project_start_row,
                        start_column=project_col,
                        end_row=row_idx - 1,
                        end_column=project_col
                    )
                current_project = proj_val
                project_start_row = row_idx

            ws.cell(row=row_idx, column=project_col, value=proj_val)
            ws.cell(row=row_idx, column=date_col, value=date_val)
            ws.cell(row=row_idx, column=content_col, value=content_val)

            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col)
                if col in style_cache:
                    copy_cell_style(style_cache[col], cell)
                else:
                    cell.font = fallback_font
                    cell.border = fallback_border
                    cell.fill = fallback_fill
                if col == date_col:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif col == content_col:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                elif col == project_col:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

            row_idx += 1

        if current_project is not None and project_start_row is not None and project_start_row < row_idx - 1:
            ws.merge_cells(
                start_row=project_start_row,
                start_column=project_col,
                end_row=row_idx - 1,
                end_column=project_col
            )

    if template_wb:
        template_wb.close()

    if default_sheet and default_sheet.title in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[default_sheet.title]

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f'SUCCESS: Performance Excel saved to {output_path}')
    return True


def cleanup_intermediate_files():
    pattern = os.path.join(OUTPUT_DIR, 'performance-*-*.xlsx')
    final_pattern = os.path.join(OUTPUT_DIR, 'performance-??????.xlsx')

    for filepath in glob.glob(pattern):
        is_final = False
        for fp in glob.glob(final_pattern):
            if os.path.abspath(filepath) == os.path.abspath(fp):
                is_final = True
                break
        if not is_final:
            try:
                os.remove(filepath)
            except OSError:
                pass


def main():
    assignee = None
    output_path = None
    list_mode = False

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--assignee' and i + 1 < len(sys.argv):
            assignee = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--output' and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--list-persons':
            list_mode = True
            i += 1
        else:
            i += 1

    col_mapping = analyze_template_columns(TEMPLATE_PATH)
    reverse_map = {v: k + 1 for k, v in col_mapping.items()}
    project_col = reverse_map.get('project', 1)
    date_col = reverse_map.get('date', 2)
    content_col = reverse_map.get('content', 3)

    if list_mode:
        all_persons = list_all_persons(project_col, date_col, content_col)
        if not all_persons:
            print('NO_DATA:No persons found')
        else:
            for name in all_persons.keys():
                print(name)
        return

    if not assignee:
        print('ERROR: --assignee is required. Use --list-persons to see available names.')
        sys.exit(1)

    if not output_path:
        output_path = os.path.join(OUTPUT_DIR, 'performance-202605.xlsx')

    if not os.path.isabs(output_path):
        output_path = os.path.abspath(output_path)

    success = generate_for_assignee(assignee, output_path)

    if success:
        cleanup_intermediate_files()


if __name__ == '__main__':
    main()
